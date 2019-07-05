import pandas as pd
import numpy as np
import keras
from keras.wrappers.scikit_learn import KerasRegressor
from keras.models import Sequential
from keras.layers import Dense, Dropout
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
import keras.optimizers
from keras.layers import LSTM
import tqdm
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
todaydf = pd.read_csv("Data/today.csv")
AAPLstocksdf = pd.read_csv("Data/AAPL.csv", parse_dates=["Date"])
# AAPLearningsdf = pd.read_csv("Data/AAPL2.csv", parse_dates=["date"])
# AAPLcleanedEDF = AAPLearningsdf.drop(["qtr"], axis=1)
AAPLdates = AAPLstocksdf["Date"]
AAPLANS = []

for index in range(len(AAPLdates)-1):
    AAPLANS.append(AAPLstocksdf.loc[index+1]["Close"])
AAPLstocksdf = AAPLstocksdf.drop(["Date"], axis=1)
train_cols = ["Open","High","Low","Close","Volume"]
df_train, df_test = train_test_split(AAPLstocksdf, test_size=0.2, shuffle=False)
x = (df_train.loc[:, train_cols]).values
min_max_scaler = MinMaxScaler()
x_train = min_max_scaler.fit_transform(x)
x_test = min_max_scaler.transform(df_test.loc[:, train_cols].values)
TIME_STEPS = 4
BATCH_SIZE = 20
def build_timeseries(mat, y_col_index):
    dim_0 = mat.shape[0]- TIME_STEPS
    dim_1 = mat.shape[1]
    x = np.zeros((dim_0, TIME_STEPS, dim_1))
    y = np.zeros((dim_0,))

    for i in tqdm.tqdm_notebook(range(dim_0)):
        x[i] = mat[i:TIME_STEPS+i]
        y[i] = mat[TIME_STEPS+i, y_col_index]
    return x,y

def trim_dataset(mat, batch_size):
    """
    trims dataset to a size that's divisible by BATCH_SIZE
    """
    no_of_rows_drop = mat.shape[0]%batch_size
    if(no_of_rows_drop > 0):
        return mat[no_of_rows_drop:]
    else:
        return mat
x_t,y_t = build_timeseries(x_train,3)
X_Test,Y_Test= build_timeseries(x_test,3)
x_t = trim_dataset(x_t, BATCH_SIZE)
y_t = trim_dataset(y_t, BATCH_SIZE)

# print(y_t)
x_temp, y_temp = build_timeseries(x_test, 3)
x_temp = trim_dataset(x_temp, BATCH_SIZE)
y_temp = trim_dataset(y_temp, BATCH_SIZE)
# x_val, x_test_t = np.split(x_temp, 2)
# y_val, y_test_t = np.split(y_temp, 2)
#AAPLstocksdf = AAPLstocksdf.drop([1257], axis=0)

# stocklist = [[MSFTstocksdf, MSFTcleanedEDF, MSFTdates, MSFTANS], [TSLAstocksdf, TSLAcleanedEDF, TSLAdates, TSLAANS], [AAPLstocksdf, AAPLcleanedEDF, AAPLdates, AAPLANS]]
# for stock in stocklist:
#     epsarray = []
#     estarray = []
#     timesince = []
#     # print(cleanedEDF.iloc[26][0] > stocksdf.iloc[2141][0])
#     ind2 = 0
#     ind3 = 0
#     ind4 = 1
#     first = True
#
#     length = len(stock[1])
#     for date in stock[0]["date"]:
#         ind3 += 1
#         timesince.append(ind3)
#         if ind2 >= length:
#             #print(length)
#             break
#         stock[3].append(abs(stock[0].loc[ind3]["close"] - stock[0].iloc[ind3]["open"]))
#         # if (stock[0].loc[ind3]["close"] - stock[0].iloc[ind3]["open"]) > 0:
#         #     stock[3].append([0,1])
#         # else:
#         #     stock[3].append([1,0])
#         epsarray.append(stock[1].iloc[ind2][2])
#         estarray.append(stock[1].iloc[ind2][1])
#         #print(date, ind2, stock[1].iloc[ind2][1], stock[1].iloc[ind2][2])
#         if date >= stock[2][ind4] and first == False:
#             # print(date, cleanedEDF.iloc[ind2][1], cleanedEDF.iloc[ind2][2])
#             # print(date, dates2[ind2], ind2)
#             del stock[2][ind4]
#             #print(date, stock[1].iloc[ind2][1], stock[1].iloc[ind2][2])
#             ind2 += 1
#             ind4 += 1
#             ind3 = 0
#         first = False
#
#     stock[0]["eps_est"] = estarray
#     stock[0]["eps"] = epsarray
#     stock[0]["time_since"] = timesince
#     #print(stock[0])
#     stock[0] = stock[0].drop(["date"], axis=1)


# merged_X = pd.concat([stocklist[2][0],stocklist[0][0], stocklist[1][0]], sort = False)
# merged_Y = stocklist[2][3] + stocklist[0][3] + stocklist[1][3]
#print(merged_Y)
#print(len(merged_Y))
# X_train, X_valid, Y_train, Y_valid = train_test_split(AAPLstocksdf,AAPLANS, train_size=0.8, test_size= 0.2, shuffle = False)


#
#
# print(X_train)
# print(Y_train)
#
# def relMeanAvErr(y_true, y_pred):
#     return keras.backend.mean((y_true - y_pred)/y_pred)
# TIME_STEPS = 3
#
# def build_timeser(mat, y_col_index)
#     dim_0 = mat.shape[0] - TIME_STEPS
#     dim_1 = mat.shape[1]
#     x = np.zeros((dim_0, TIME_STEPS, dim_1))
#     y = np.zeros((dim_0,))
#
#     for i in range(dim_0):
#         x[i] = mat[i:TIME_STEPS + i]
#         y[i] = mat[TIME_STEPS + i, y_col_index]
#     print("length of time-series i/o", x.shape, y.shape)
#
#
#     return x, y
#
#
def basemodel():
    model = Sequential()
    model.add(LSTM(100, batch_input_shape=(BATCH_SIZE, TIME_STEPS, x_t.shape[2]), dropout=0.0, recurrent_dropout=0.0,
                   stateful=True, kernel_initializer='random_uniform'))
    model.add(Dense(100, kernel_initializer='normal', activation='relu'))
    model.add(Dense(100, kernel_initializer='normal', activation= 'relu'))
    model.add(Dense(100, kernel_initializer='normal', activation='relu'))
    model.add(Dense(1, kernel_initializer='normal', activation='linear'))
    model.compile(loss='mse', optimizer='adam', metrics=['mean_absolute_error', 'mape'])
    return model

#
#
predictor = KerasRegressor(build_fn=basemodel, epochs = 100, verbose =2)

predictor.fit(x_t, y_t ,epochs=100, batch_size = BATCH_SIZE, )
# kfold = KFold(random_state=1, n_splits=10)
# estimators = []
# estimators.append(('standardize', StandardScaler()))
# estimators.append(('mlp', predictor))
# pipeline = Pipeline(estimators)
# #results = cross_val_score(pipeline, X_train, Y_train, cv=kfold)
# pipeline.fit(X_train,Y_train)
# #print("Results: %.2f (%.2f) MSE" % (results.mean(), results.std()))
from sklearn.metrics import mean_absolute_error
prediction = predictor.predict(x_temp, batch_size = BATCH_SIZE)
mae = mean_absolute_error(y_temp, prediction)
mape = keras.metrics.mean_absolute_percentage_error(y_temp, prediction)
# #td = keras.backend.std(prediction)
#
#
# tompredict = predictor.predict(todaydf)
ind = 0
prediction = prediction.reshape(-1,1)
y_val = y_temp.reshape(-1,1)
prediction =(prediction * min_max_scaler.data_range_[3]) + min_max_scaler.data_min_[3]#min_max_scaler.inverse_transform(prediction)
y_val =(y_val * min_max_scaler.data_range_[3]) + min_max_scaler.data_min_[3]#min_max_scaler.inverse_transform(y_val)

singlearraypred = []
singlearrayy_val = []
for val in prediction:
    singlearraypred.append(prediction[ind][0])
    singlearrayy_val.append(y_val[ind][0])
    print(singlearraypred[ind], singlearrayy_val[ind])
    ind+=1

# x = range(len(prediction))
# print("tom: ", tompredict)
print("mae: " , mae, "mape: ", mape)
# # # error = []
# # # for u in x:
# # #     error.append(abs(prediction[u]-Y_valid[u]))
# # #print('std: ', np.std(error))
import openpyxl as xl
#
wb = xl.load_workbook("Data/output.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
sheet.cell(1,1).value = "prediction"
sheet.cell(1,2).value = "Y_valid"
for z in range(len(prediction)):
    sheet.cell(z+2, 1).value = singlearraypred[z]
    sheet.cell(z+2, 2).value= singlearrayy_val[z]
wb.save("Data/output.xlsx")