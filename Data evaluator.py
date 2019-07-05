import openpyxl as xl
import pandas as pd
import numpy as np
import keras
from keras.wrappers.scikit_learn import KerasRegressor
from keras.models import Sequential
from keras.layers import Dense, Dropout
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
import keras.optimizers
from keras.layers import LSTM
import tqdm
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler

from sklearn.metrics import mean_absolute_error
prediction = predictor.predict(x_temp, batch_size = BATCH_SIZE)
mae = mean_absolute_error(y_temp, prediction)
mape = keras.metrics.mean_absolute_percentage_error(y_temp, prediction)
ind = 0
prediction = prediction.reshape(-1,1)
y_val = y_temp.reshape(-1,1)
prediction =(prediction * min_max_scaler.data_range_[3]) + min_max_scaler.data_min_[3]#min_max_scaler.inverse_transform(prediction)
y_val =(y_val * min_max_scaler.data_range_[3]) + min_max_scaler.data_min_[3]#min_max_scaler.inverse_transform(y_val)

singlearraypred = []
singlearrayy_val = []
for val in prediction:
    singlearraypred.append(prediction[ind][0])
    singlearrayy_val.append(y_val[ind][0])
    print(singlearraypred[ind], singlearrayy_val[ind])
    ind+=1

print("mae: " , mae, "mape: ", mape)



wb = xl.load_workbook("Data/output.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
sheet.cell(1,1).value = "prediction"
sheet.cell(1,2).value = "Y_valid"
for z in range(len(prediction)):
    sheet.cell(z+2, 1).value = singlearraypred[z]
    sheet.cell(z+2, 2).value= singlearrayy_val[z]
wb.save("Data/output.xlsx")